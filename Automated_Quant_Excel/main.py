# ============================================================
# FINANCIAL PDF → EXCEL PIPELINE
# ABBYY Cloud OCR SDK + Gemini API (google-genai SDK)
# ============================================================
# HOW TO CONNECT ABBYY CLOUD OCR SDK:
# 1. Go to https://cloud.ocrsdk.com and sign in / register
# 2. Create a new Application → note Application ID + Password
# 3. In your app details, note the "Processing Location" URL
#    e.g. https://cloud-westus.ocrsdk.com
# 4. Paste credentials and the FULL location URL below
#
# HOW TO GET GEMINI API KEY:
# 1. Go to https://aistudio.google.com/app/apikey
# 2. Create API key → paste below
#
# INSTALL DEPENDENCIES:
# pip3 install requests openpyxl google-genai
# ============================================================

import requests
import time
import json
import datetime
import xml.etree.ElementTree as ET
from google import genai
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ============================================================
# CONFIGURATION — FILL THESE IN
# ============================================================
# ABBYY_APP_ID       = "2bb07070-ed65-458a-8d16-30c5a28a822a"
# ABBYY_APP_PASSWORD = "F54fhXD9+WoQ1OO6PPb6iHMf"

# ⚠️  IMPORTANT: Use the region-specific URL from your ABBYY app details page.
# NOT https://cloud.ocrsdk.com — that's only the portal login.
# Examples: https://cloud-westus.ocrsdk.com  |  https://cloud-eu.ocrsdk.com
# ABBYY_SERVICE_URL  = "https://cloud-westus.ocrsdk.com"

# GEMINI_API_KEY     = "AIzaSyCK4AO3TDOIsqZImB36nKH41PLxKmnKX7k"
# GEMINI_MODEL       = "gemini-2.5-flash"

# Path to your PDF
PDF_PATH           = str(Path.home() / "Downloads" / "xyz.pdf")

# Output Excel path
OUTPUT_EXCEL       = str(Path.home() / "Downloads" / "xyz_financial_output_updated.xlsx")

# For year-on-year update mode — set path to existing Excel, else leave as None
EXISTING_EXCEL   = str(Path.home() / "Downloads" / "xyz_financial_output.xlsx")
# EXISTING_EXCEL     = None

# Company name (used in Gemini prompt)
COMPANY_NAME       = "XYZ"

# ============================================================
# EXCEL SCHEMA RULES — passed to Gemini
# ============================================================
EXCEL_RULES = """
OUTPUT FORMAT: Flat table — one row per metric. One sheet per financial table.
Fixed columns (in this exact order):
primary_key | date_last_updated | doc_page_num | file_page_num | table_id |
geo_1_id | geo_1_name | geo_1_type | geo_2_id | geo_2_name | geo_2_type |
dim_4_id | dim_4_name | dim_3_id | dim_3_name | dim_2_id | dim_2_name |
dim_1_id | dim_1_name | metric_id | metric_name | source_metric_id |
source_metric_name | indentation | process_flag | base_factor |
display_power_factor | data_frequency | aggregation_method | unit | unit_type |
note_id | note_reference | cumulative_periods | comments |
check_sum | concat | formula | [YEAR COLUMNS — only years present in the PDF table]

RULES:
1. Extract EVERY row — even if ALL year columns are blank (dashes or empty)
2. Values: if "-" or blank → leave year column EMPTY (not 0, not "-", just empty string)
3. Convert parentheses to negative: (123) → -123, (1,234) → -1234
4. metric_name = standardized row label:
   - Remove special chars: . , ( ) * / - _ → replace with blank space
   - Replace % → " percentage "
   - Replace + → " Plus "
   - Collapse multiple spaces → single space
   - Trim leading/trailing spaces
5. source_metric_name = EXACT original text from PDF — no modification
6. Leave ALL *_id columns EMPTY always (metric_id, source_metric_id, geo_*_id, dim_*_id)
7. process_flag = 1 always
8. cumulative_periods = true always
9. comments = false by default
   comments = true ONLY when metric_name contains "Total" OR metric is a calculated sum/subtotal
10. geo_1_id, geo_1_name, geo_1_type, geo_2_id, geo_2_name, geo_2_type = ALWAYS EMPTY
11. data_frequency = ALWAYS EMPTY
12. aggregation_method = ALWAYS EMPTY
13. published_date = ALWAYS EMPTY (not a column)
14. reported_date = ALWAYS EMPTY (not a column)

15. base_factor, unit, unit_type — DETECT FROM PDF:
    - Look for text near/above each table like "in thousands", "USD millions", "₹ in Crores", "% of revenue"
    - base_factor: if "thousands" → 1000, if "millions" → 1000000, if "crores" → 10000000, if "lakhs" → 100000, else → 1
    - display_power_factor: if base_factor=1000 → 3, if 1000000 → 6, if 10000000 → 7, if 100000 → 5, else → 0
    - unit: detect currency symbol or text → "usd", "eur", "inr", "gbp", "pln" etc. If percentage table → "percentage". If no currency → ""
    - unit_type: if currency → "ccy", if percentage → "percentage", if ratio → "ratio", if count → "quantity", else → ""
    - If nothing mentioned → base_factor=1, display_power_factor=0, unit="", unit_type=""

16. YEAR COLUMNS — CRITICAL:
    - Only create columns for years ACTUALLY present in that specific table
    - If table has 2024 and 2025 → only "2024" and "2025" columns
    - If table has 2022, 2023, 2024 → only those 3 columns
    - Do NOT add years that are not in the table

17. table_id — DETECT FROM PDF:
    - Use the table's title/heading as the table_id (snake_case, lowercase)
    - e.g. "Consolidated Balance Sheet" → "consolidated_balance_sheet"
    - e.g. "Cash and Cash Equivalents" → "cash_and_cash_equivalents"
    - e.g. "Statement of Profit and Loss" → "statement_of_profit_and_loss"

18. Dimension hierarchy (section headers → dim columns):
    - dim_1_name = main section header (e.g. "Current Assets", "Non-Current Liabilities")
    - dim_2_name = sub-section under dim_1
    - dim_3_name = sub-section under dim_2
    - dim_4_name = sub-section under dim_3
    - Apply same dim values to ALL metrics under that section until next section starts
    - Section headers themselves → output as a row with empty year values
    - Apply dim names to all sub-metrics below them

19. indentation = numeric depth level of the metric (0 = top level, 1 = one indent, 2 = two indents, etc.)
"""

# ============================================================
# GEMINI HELPER
# ============================================================
def call_gemini(prompt: str) -> str:
    """Call Gemini API and return clean JSON text."""
    import urllib.request, urllib.error
    url  = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
    body = json.dumps({"contents": [{"parts": [{"text": prompt}]}]}).encode()
    req  = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(req) as resp:
            result = json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        raise Exception(f"Gemini error {e.code}: {e.read().decode()}")
    raw = result["candidates"][0]["content"]["parts"][0]["text"].strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"): raw = raw[4:]
    return raw.strip()

# ============================================================
# STEP 1: ABBYY CLOUD OCR — SUBMIT PDF
# Supported exportFormat values: txt, xml, docx, xlsx, rtf, pptx, pdf, pdfa
# "json" is NOT a valid ABBYY export format — we use "xml" and parse it
# ============================================================
def abbyy_submit_pdf(pdf_path: str) -> str:
    """Upload PDF to ABBYY Cloud OCR. Returns task ID."""
    print(f"[ABBYY] Submitting PDF: {pdf_path}")

    url    = f"{ABBYY_SERVICE_URL}/processImage"
    params = {
        "exportFormat": "xml",          # xml gives full structured text output
        "language":     "English",
        "profile":      "documentArchiving",
    }

    with open(pdf_path, "rb") as f:
        response = requests.post(
            url,
            params  = params,
            files   = {"file": (Path(pdf_path).name, f, "application/pdf")},
            auth    = (ABBYY_APP_ID, ABBYY_APP_PASSWORD)
        )

    if response.status_code != 200:
        raise Exception(f"ABBYY submit failed: {response.status_code} — {response.text}")

    # ABBYY returns XML: <response><task id="..." status="Queued" .../></response>
    root    = ET.fromstring(response.text)
    task    = root.find("task")          # direct child named "task"
    if task is None:
        task = root                      # fallback: root itself carries attribs
    task_id = task.attrib.get("id")
    if not task_id:
        raise Exception(f"ABBYY: could not find task ID in response:\n{response.text}")

    print(f"[ABBYY] Task ID: {task_id}")
    return task_id

# ============================================================
# STEP 2: ABBYY — POLL UNTIL DONE, DOWNLOAD XML, RETURN AS TEXT
# ABBYY says: poll every 2–3 seconds, no auth needed for resultUrl download
# ============================================================
def abbyy_wait_for_result(task_id: str, poll_interval: int = 3) -> str:
    """
    Poll ABBYY until task completes.
    Downloads the XML result and converts it to plain text for Gemini.
    Returns: plain text string of the full document.
    """
    print(f"[ABBYY] Waiting for task {task_id}...")

    while True:
        response = requests.get(
            f"{ABBYY_SERVICE_URL}/getTaskStatus",
            params = {"taskId": task_id},
            auth   = (ABBYY_APP_ID, ABBYY_APP_PASSWORD)
        )

        if response.status_code != 200:
            raise Exception(f"ABBYY status check failed: {response.status_code} — {response.text}")

        root   = ET.fromstring(response.text)
        task   = root.find("task")       # <response><task .../></response>
        if task is None:
            task = root
        status = task.attrib.get("status", "")
        print(f"[ABBYY] Status: {status}")

        if status == "Completed":
            result_url = task.attrib.get("resultUrl")
            if not result_url:
                raise Exception("ABBYY completed but no resultUrl in response.")
            print("[ABBYY] Done! Downloading OCR result...")

            # No auth headers for result download — ABBYY requirement
            xml_response = requests.get(result_url)
            xml_response.raise_for_status()

            # Parse ABBYY XML → extract all text blocks as plain text
            ocr_text = _abbyy_xml_to_text(xml_response.text)
            print(f"[ABBYY] Extracted {len(ocr_text)} characters of text.")
            return ocr_text

        if status in ("ProcessingFailed", "NotEnoughCredits"):
            raise Exception(f"ABBYY processing failed with status: {status}")

        time.sleep(poll_interval)

def _abbyy_xml_to_text(xml_string: str) -> str:
    """
    Parse ABBYY output XML and extract all text in reading order.
    ABBYY XML structure: document → page → block → text → par → line → formatting → charParams
    We collect all <line> text joined by newlines, blocks separated by double newlines.
    """
    try:
        root = ET.fromstring(xml_string)
    except ET.ParseError:
        # If XML is malformed, return raw text stripped of tags as fallback
        import re
        return re.sub(r"<[^>]+>", " ", xml_string)

    ns = ""
    # Detect namespace if present
    if root.tag.startswith("{"):
        ns = root.tag.split("}")[0] + "}"

    pages_text = []
    page_num   = 0

    for page in root.iter(f"{ns}page"):
        page_num += 1
        page_lines = [f"\n--- PAGE {page_num} ---"]

        for block in page.iter(f"{ns}block"):
            block_lines = []
            for line in block.iter(f"{ns}line"):
                # Collect all charParams text in this line
                chars = "".join(
                    (cp.text or "")
                    for cp in line.iter(f"{ns}charParams")
                )
                if chars.strip():
                    block_lines.append(chars.strip())
            if block_lines:
                page_lines.append("\n".join(block_lines))
                page_lines.append("")  # blank line between blocks

        pages_text.append("\n".join(page_lines))

    return "\n".join(pages_text)

# ============================================================
# STEP 3A: GEMINI — FRESH PDF
# ============================================================
def gemini_extract_fresh(ocr_text: str) -> dict:
    """
    Send extracted OCR text to Gemini.
    Gemini identifies tables, assigns IDs, detects statement types,
    year columns, base_factor/unit/unit_type, and returns flat row schema.
    """
    print("[GEMINI] Fresh extraction mode...")

    prompt = f"""
You are a financial document analysis expert.

Company: {COMPANY_NAME}

FULL DOCUMENT TEXT (extracted via OCR from financial PDF):
{ocr_text[:50000]}

Your tasks — follow ALL rules below EXACTLY:
{EXCEL_RULES}

Return ONLY valid JSON in EXACTLY this format. No text outside JSON.
The "year_columns" array must contain ONLY the years actually present in that table.
The "rows" must have keys for each year in year_columns.

{{
  "company": "{COMPANY_NAME}",
  "tables": [
    {{
      "table_id": "consolidated_balance_sheet",
      "statement_type": "BALANCE_SHEET",
      "page_number": 5,
      "year_columns": ["2024", "2025"],
      "base_factor": 1000,
      "display_power_factor": 3,
      "unit": "inr",
      "unit_type": "ccy",
      "rows": [
        {{
          "primary_key": "",
          "date_last_updated": "",
          "doc_page_num": 5,
          "file_page_num": 5,
          "table_id": "consolidated_balance_sheet",
          "geo_1_id": "", "geo_1_name": "", "geo_1_type": "",
          "geo_2_id": "", "geo_2_name": "", "geo_2_type": "",
          "dim_4_id": "", "dim_4_name": "",
          "dim_3_id": "", "dim_3_name": "",
          "dim_2_id": "", "dim_2_name": "",
          "dim_1_id": "", "dim_1_name": "Current Assets",
          "metric_id": "",
          "metric_name": "Cash and Cash Equivalents",
          "source_metric_id": "",
          "source_metric_name": "Cash & Cash Equivalents",
          "indentation": 1,
          "process_flag": 1,
          "base_factor": 1000,
          "display_power_factor": 3,
          "data_frequency": "",
          "aggregation_method": "",
          "unit": "inr",
          "unit_type": "ccy",
          "note_id": "", "note_reference": "",
          "cumulative_periods": true,
          "comments": false,
          "check_sum": "", "concat": "", "formula": "",
          "2024": 50000,
          "2025": 62000
        }}
      ]
    }}
  ]
}}
"""
    raw = call_gemini(prompt)
    return json.loads(raw)

# ============================================================
# STEP 3B: GEMINI — YEAR-ON-YEAR UPDATE MODE
# ============================================================
# USE THIS instead of gemini_extract_fresh() when updating an
# existing Excel with a new year's PDF data.
#
# TO ACTIVATE:
# 1. Uncomment the function below
# 2. Set EXISTING_EXCEL = "path/to/your/previous.xlsx" above
# 3. The pipeline auto-switches when EXISTING_EXCEL is set
#
def gemini_extract_update(ocr_text: str, existing_excel_path: str) -> dict:
    """
    YoY Update:
    - Reads existing Excel (previous years)
    - Matches metrics from new PDF intelligently (handles renamed metrics)
    - Adds new year as a new column only (no extra columns added)
    - Adds new rows for new metrics, keeps removed metrics with empty new-year value
    - Preserves ALL existing Table IDs and historical year columns untouched
    """
    print("[GEMINI] Year-on-year update mode...")
    wb_existing   = load_workbook(existing_excel_path)
    existing_data = {}
    for sheet_name in wb_existing.sheetnames:
        ws = wb_existing[sheet_name]
        existing_data[sheet_name] = [
            [cell.value for cell in row] for row in ws.iter_rows()
        ]
    prompt = f"""
You are a financial document analysis expert specializing in year-on-year comparisons.
Company: {COMPANY_NAME}
EXISTING EXCEL DATA (historical — DO NOT modify these columns or values):
{json.dumps(existing_data, indent=2)[:20000]}
NEW OCR TEXT (from new year PDF):
{ocr_text[:30000]}
Rules: {EXCEL_RULES}
Additional YoY rules:
- Detect which new year is in the new PDF (e.g. 2025)
- Add ONLY that new year as a new column
- Match metrics intelligently (e.g. "Net Revenue" == "Revenue from Operations" → same row)
- New metrics in new PDF → add as new rows
- Metrics removed in new PDF → keep row, leave new year column empty
- Preserve ALL existing Table IDs — only new tables get new IDs
- base_factor, unit, unit_type — re-detect from new PDF context (may have changed)
Return ONLY valid JSON in same format as fresh extraction. No text outside JSON.
"""
    raw = call_gemini(prompt)
    return json.loads(raw)

# ============================================================
# CONVERTER — Flat sheet-keyed JSON → tables[] format
# Gemini YoY mode mirrors the existing Excel structure back,
# so we need to convert it to the standard tables[] format.
# ============================================================
def convert_flat_to_tables(flat_data: dict) -> dict:
    tables = []

    for sheet_name, sheet_rows in flat_data.items():
        if sheet_name == "SUMMARY" or not sheet_rows:
            continue

        headers = [str(h) if h is not None else "" for h in sheet_rows[0]]
        year_cols = [h for h in headers if h.isdigit() and len(h) == 4]

        first_row = sheet_rows[1] if len(sheet_rows) > 1 else []
        row_dict  = dict(zip(headers, first_row))

        table_id    = row_dict.get("table_id", sheet_name)
        page_num    = row_dict.get("doc_page_num", 0) or 0
        base_factor = row_dict.get("base_factor", 1) or 1
        disp_power  = row_dict.get("display_power_factor", 0) or 0
        unit        = row_dict.get("unit", "") or ""
        unit_type   = row_dict.get("unit_type", "") or ""

        tid = str(table_id).lower()
        if "balance" in tid or "financial_position" in tid:
            stmt_type = "BALANCE_SHEET"
        elif "profit" in tid or "loss" in tid or "income" in tid:
            stmt_type = "INCOME_STATEMENT"
        elif "cash" in tid:
            stmt_type = "CASH_FLOW"
        elif "equity" in tid:
            stmt_type = "EQUITY"
        else:
            stmt_type = "UNKNOWN"

        rows = []
        for raw_row in sheet_rows[1:]:
            row = dict(zip(headers, raw_row))
            row = {k: (v if v is not None else "") for k, v in row.items()}
            rows.append(row)

        tables.append({
            "table_id":             table_id,
            "statement_type":       stmt_type,
            "page_number":          int(page_num),
            "year_columns":         year_cols,
            "base_factor":          base_factor,
            "display_power_factor": disp_power,
            "unit":                 unit,
            "unit_type":            unit_type,
            "rows":                 rows,
        })

    return {"company": COMPANY_NAME, "tables": tables}

# ============================================================
# STEP 4: WRITE TO EXCEL
# ============================================================
def write_excel(structured_data: dict, output_path: str):
    """Write flat schema to Excel — one sheet per table, one row per metric."""
    print(f"[EXCEL] Writing to {output_path}...")

    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")
    LABEL_FONT   = Font(bold=True, size=10)
    NEG_FONT     = Font(color="FF0000", size=10)
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT        = Alignment(horizontal="right",  vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")
    thin_side    = Side(style="thin", color="D0D0D0")
    THIN_BORDER  = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    today        = datetime.date.today().isoformat()
    tables       = structured_data.get("tables", [])
    summary_rows = []

    FIXED_COLS = [
        "primary_key", "date_last_updated", "doc_page_num", "file_page_num", "table_id",
        "geo_1_id", "geo_1_name", "geo_1_type", "geo_2_id", "geo_2_name", "geo_2_type",
        "dim_4_id", "dim_4_name", "dim_3_id", "dim_3_name", "dim_2_id", "dim_2_name",
        "dim_1_id", "dim_1_name", "metric_id", "metric_name", "source_metric_id",
        "source_metric_name", "indentation", "process_flag", "base_factor",
        "display_power_factor", "data_frequency", "aggregation_method",
        "unit", "unit_type", "note_id", "note_reference", "cumulative_periods",
        "comments", "check_sum", "concat", "formula"
    ]

    for table in tables:
        table_id   = table["table_id"]
        stmt_type  = table["statement_type"]
        page_num   = table["page_number"]
        year_cols  = [str(y) for y in table.get("year_columns", [])]
        rows       = table["rows"]

        ALL_COLS   = FIXED_COLS + year_cols
        sheet_name = f"{table_id}"[:31]
        ws         = wb.create_sheet(title=sheet_name)

        # Row 1 — Header
        for ci, col in enumerate(ALL_COLS, 1):
            cell           = ws.cell(row=1, column=ci, value=col)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER
            cell.border    = THIN_BORDER

        # Row 2+ — Data
        for ri, row in enumerate(rows, 2):
            row["date_last_updated"] = today
            fill = ALT_FILL if ri % 2 == 0 else PatternFill()

            for ci, col in enumerate(ALL_COLS, 1):
                val  = row.get(col, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill   = fill
                cell.border = THIN_BORDER

                if col == "metric_name":
                    cell.font      = LABEL_FONT
                    cell.alignment = LEFT
                elif col in year_cols:
                    cell.alignment = RIGHT
                    try:
                        num = float(str(val).replace(",", "")) if val != "" else None
                        if num is not None and num < 0:
                            cell.font = NEG_FONT
                        if num is not None:
                            cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        pass
                else:
                    cell.alignment = LEFT

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

        summary_rows.append({
            "table_id":    table_id,
            "type":        stmt_type,
            "page":        page_num,
            "years":       ", ".join(year_cols),
            "sheet_name":  sheet_name,
            "row_count":   len(rows),
            "base_factor": table.get("base_factor", 1),
            "unit":        table.get("unit", ""),
            "unit_type":   table.get("unit_type", ""),
        })

    # Summary sheet
    ws_s = wb.create_sheet(title="SUMMARY", index=0)
    ws_s["A1"] = f"Financial Report — {structured_data.get('company', '')} | Extracted: {today}"
    ws_s["A1"].font = Font(bold=True, size=13, color="1F3864")

    sum_headers = ["Table ID", "Statement Type", "Page", "Years", "Sheet Name", "Row Count", "Base Factor", "Unit", "Unit Type"]
    for ci, h in enumerate(sum_headers, 1):
        cell           = ws_s.cell(row=3, column=ci, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER

    for ri, row in enumerate(summary_rows, 4):
        vals = [row["table_id"], row["type"], row["page"], row["years"],
                row["sheet_name"], row["row_count"], row["base_factor"], row["unit"], row["unit_type"]]
        fill = ALT_FILL if ri % 2 == 0 else PatternFill()
        for ci, val in enumerate(vals, 1):
            cell        = ws_s.cell(row=ri, column=ci, value=val)
            cell.fill   = fill
            cell.border = THIN_BORDER

    for col in ws_s.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws_s.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    ws_s.freeze_panes = "A4"
    wb.save(output_path)
    print(f"[EXCEL] Saved → {output_path}")

# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("FINANCIAL PDF → EXCEL PIPELINE")
    print("=" * 60)


    # Step 1 — Upload PDF to ABBYY
    task_id  = abbyy_submit_pdf(PDF_PATH)

    # Step 2 — Poll until done, get plain text
    ocr_text = abbyy_wait_for_result(task_id)

    # Step 3 — Gemini extracts structured data
    if EXISTING_EXCEL is None:
        structured_data = gemini_extract_fresh(ocr_text)
    else:
        # YoY mode — uncomment gemini_extract_update() above first
        row_yoy_data = gemini_extract_update(ocr_text, EXISTING_EXCEL)
        structured_data = convert_flat_to_tables(row_yoy_data)
        # raise NotImplementedError("Uncomment gemini_extract_update() above to use YoY mode")

    # Save raw Gemini JSON for debugging
    debug_path = OUTPUT_EXCEL.replace(".xlsx", "_gemini_raw.json")
    with open(debug_path, "w") as f:
        json.dump(structured_data, f, indent=2)
    print(f"[DEBUG] Raw Gemini JSON → {debug_path}")

    # Step 4 — Write to Excel
    write_excel(structured_data, OUTPUT_EXCEL)

    print("=" * 60)
    print(f"DONE! → {OUTPUT_EXCEL}")
    print("=" * 60)

if __name__ == "__main__":
    main()
