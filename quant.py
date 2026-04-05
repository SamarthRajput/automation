import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os
import re
import asyncio
import concurrent.futures
from collections import defaultdict
from io import BytesIO
from datetime import datetime, timedelta, timezone
from azure.storage.blob import ContainerClient
from azure.core.exceptions import ResourceNotFoundError
import warnings
warnings.filterwarnings('ignore')


# ═══════════════════════════════════════════════════════════════
#  ▶▶  CONFIGURE HERE  ◀◀
# ═══════════════════════════════════════════════════════════════
CONNECTION_STRING = (
    "xx"
    "yy"
    "zz"
)
CONTAINER_NAME = "sgfr01intern"

# ── Transition Sheet (LOCAL hardcoded path — NOT from blob) ──
TRANSITION_LOCAL_PATH = r"C:\Users\SamarthRajput\Downloads\Transition_Sheet_06032026.xlsx"
TRANSITION_SHEET_NAME = "Transformation Transition"

# ── Output (local) ───────────────────────────────────────────
OUTPUT_PATH     = r"C:\Users\SamarthRajput\Downloads\quant_ri_error_report.xlsx"
STALE_FILE_DAYS = 181
# ═══════════════════════════════════════════════════════════════
# company_id and granularity are read from the local Transition
# Sheet — NOT hardcoded. Every unique (company_id, granularity)
# pair is processed automatically.
# ═══════════════════════════════════════════════════════════════


NON_YEAR_COLS = {
    'primary_key', 'date_last_updated', 'published_date', 'reported_date',
    'doc_page_num', 'file_page_num', 'table_id', 'country',
    'geo_1_id', 'geo_1_name', 'geo_1_type',
    'geo_2_id', 'geo_2_name', 'geo_2_type',
    'dim_4_id', 'dim_4_name', 'dim_3_id', 'dim_3_name',
    'dim_2_id', 'dim_2_name', 'dim_1_id', 'dim_1_name',
    'metric_id', 'metric_name', 'source_metric_id', 'source_metric_name',
    'indentation', 'process_flag', 'base_factor', 'display_power_factor',
    'data_frequency', 'aggregation_method', 'unit', 'unit_type',
    'note_id', 'note_reference', 'cumulative_periods', 'comments',
    'check_sum', 'concat', 'formula',
}

YEAR_PATTERN      = re.compile(r'^\d{4}$')
PF0_STATUS_VALUES = {'pf0', 'pf_0', 'pf-0', 'pf 0'}

COUNTED_ERROR_TYPES = [
    'change_in_page_number',
    'change_in_values',
    'missing_table_in_quant',
    'change_in_dim_names',
    'change_in_unit',
    'change_in_unit_type',
    'change_in_note_id',
    'change_in_comments',
    'change_in_indentation',
]

MISSING_ERROR_TYPES = [
    'missing_table_in_both',
    'missing_table_in_both_pf0',
    'missing_table_in_initial',
]

ERROR_TYPES  = COUNTED_ERROR_TYPES + MISSING_ERROR_TYPES
ERROR_LABELS = {
    'change_in_page_number':      'Page Number Change',
    'change_in_values':           'Value Change',
    'missing_table_in_quant':     'Missing in Quant File',
    'missing_table_in_initial':   'Missing in Initial File',
    'missing_table_in_both':      'Missing in Both Files',
    'missing_table_in_both_pf0':  'Missing in Both Files (PF0)',
    'change_in_dim_names':        'Dim Name Change',
    'change_in_unit':             'Unit Change',
    'change_in_unit_type':        'Unit Type Change',
    'change_in_note_id':          'Note ID Change',
    'change_in_comments':         'Comments Change',
    'change_in_indentation':      'Indentation Change',
}

MISSING_TYPE_LABELS = {
    'missing_table_in_both':      'Absent from Both Files',
    'missing_table_in_both_pf0':  'Absent from Both Files (PF0)',
    'missing_table_in_initial':   'Present in Quant Only (Not in Initial)',
}

# ── Styles ───────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', start_color='1F3864')
SUB_FILL   = PatternFill('solid', start_color='2E75B6')
ALT_FILL   = PatternFill('solid', start_color='D6E4F0')
ERR_FILL   = PatternFill('solid', start_color='FFE0E0')
OK_FILL    = PatternFill('solid', start_color='E2EFDA')
WARN_FILL  = PatternFill('solid', start_color='FFF2CC')
WHITE_FILL = PatternFill('solid', start_color='FFFFFF')

HDR_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY_FONT  = Font(name='Arial', size=9)
BOLD_FONT  = Font(name='Arial', bold=True, size=9)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F3864')

thin   = Side(style='thin', color='B0B0B0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)


# ════════════════════════════════════════════════════════════════
# AZURE CORE HELPERS
# ════════════════════════════════════════════════════════════════

def get_container_client() -> ContainerClient:
    return ContainerClient.from_connection_string(
        conn_str=CONNECTION_STRING,
        container_name=CONTAINER_NAME,
    )


def download_blob_to_bytesio_sync(blob_path: str) -> BytesIO:
    cc     = get_container_client()
    client = cc.get_blob_client(blob_path)
    stream = client.download_blob()
    return BytesIO(stream.readall())


def read_blob_as_df(blob_path: str, sheet_name=None) -> pd.DataFrame:
    bio = download_blob_to_bytesio_sync(blob_path)
    ext = os.path.splitext(blob_path)[1].lower()
    if ext == '.csv':
        df = pd.read_csv(bio, dtype=str, keep_default_na=False)
    else:
        kw = {'dtype': str, 'keep_default_na': False}
        if sheet_name:
            kw['sheet_name'] = sheet_name
        df = pd.read_excel(bio, **kw)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
    return df


def is_blob_stale(last_modified, max_age_days: int = STALE_FILE_DAYS) -> bool:
    if last_modified is None:
        return True
    now_utc = datetime.now(timezone.utc)
    cutoff  = now_utc - timedelta(days=max_age_days)
    if last_modified.tzinfo is None:
        last_modified = last_modified.replace(tzinfo=timezone.utc)
    return last_modified < cutoff


# ════════════════════════════════════════════════════════════════
# BLOB PATH BUILDERS
# ════════════════════════════════════════════════════════════════

def get_final_blob_path(company_id: str, granularity: str) -> str:
    base = (
        f"transformation/processed/unstructured-output/"
        f"quantitative-output/{company_id}/"
    )
    if str(granularity).strip().lower() == 'quarter':
        base += 'quarter/'
    return f"{base}{company_id}_quantitative.csv"


def get_initial_bkp_prefix(company_id: str, granularity: str) -> str:
    base = (
        f"transformation/working-files/unstructured_data/"
        f"quantitative_checks/{company_id}/"
    )
    if str(granularity).strip().lower() == 'quarter':
        base += 'quarter/'
    return base


# ════════════════════════════════════════════════════════════════
# TRANSITION SHEET — LOCAL FILE READER
# ════════════════════════════════════════════════════════════════

def load_transition_sheet_local() -> pd.DataFrame:
    """
    Read the Transition Sheet from the hardcoded LOCAL path.
    Does NOT touch Azure for this file.
    """
    path = TRANSITION_LOCAL_PATH
    print(f"  → Loading Transition Sheet (local): {path}")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Transition Sheet not found at: {path}")

    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        df = pd.read_csv(path, dtype=str, keep_default_na=False)
    else:
        df = pd.read_excel(path, sheet_name=TRANSITION_SHEET_NAME,
                           dtype=str, keep_default_na=False)

    df.columns = [str(c).strip() for c in df.columns]
    df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
    print(f"  → Loaded {len(df)} rows from local Transition Sheet")
    return df


# ════════════════════════════════════════════════════════════════
# AZURE BLOB FILE LOADERS
# ════════════════════════════════════════════════════════════════

def load_final_file_sync(company_id: str, granularity: str) -> pd.DataFrame:
    blob_path = get_final_blob_path(company_id, granularity)
    print(f"    [final]   {blob_path}")
    df = read_blob_as_df(blob_path)
    if 'table_id' in df.columns:
        df['table_id'] = df['table_id'].astype(str).str.strip()
    return df


def load_initial_file_sync(company_id: str, granularity: str) -> tuple:
    prefix = get_initial_bkp_prefix(company_id, granularity)
    cc     = get_container_client()

    latest_blob = None
    latest_time = None

    for blob in cc.list_blobs(name_starts_with=prefix):
        name = blob.name.lower()
        if '_bkp' in name and (name.endswith('.csv') or name.endswith('.xlsx')):
            if latest_time is None or blob.last_modified > latest_time:
                latest_time = blob.last_modified
                latest_blob = blob.name

    if not latest_blob:
        raise ResourceNotFoundError(
            f"No _bkp file found under prefix '{prefix}' "
            f"for company='{company_id}', granularity='{granularity}'"
        )

    print(f"    [initial] {latest_blob}  (last_modified={latest_time})")
    df = read_blob_as_df(latest_blob)
    if 'table_id' in df.columns:
        df['table_id'] = df['table_id'].astype(str).str.strip()
    return df, latest_time


async def load_final_file_async(company_id: str, granularity: str) -> pd.DataFrame:
    return await asyncio.to_thread(load_final_file_sync, company_id, granularity)


async def load_initial_file_async(company_id: str, granularity: str) -> tuple:
    return await asyncio.to_thread(load_initial_file_sync, company_id, granularity)


# ════════════════════════════════════════════════════════════════
# TRANSITION SHEET UTILITIES
# ════════════════════════════════════════════════════════════════

def normalise_col(df, candidates):
    col_map = {c.lower().replace(' ', '').replace('_', ''): c for c in df.columns}
    for cand in candidates:
        key = cand.lower().replace(' ', '').replace('_', '')
        if key in col_map:
            return col_map[key]
    return None


def extract_company_granularity_pairs(df_trans: pd.DataFrame) -> list:
    """
    Read every unique (company_id, granularity) pair from the local Transition Sheet.
    Granularity is derived from the period/data_frequency column:
      'quarter' → 'quarter', everything else → 'annual'.
    Falls back to 'annual' if no granularity column is found.
    """
    company_col = normalise_col(df_trans, ['company_id', 'companyid', 'key'])
    gran_col    = normalise_col(df_trans, ['granularity', 'datafrequency', 'data_frequency',
                                            'frequency', 'period', 'period_type', 'annual'])

    if company_col is None:
        raise ValueError("Cannot locate 'company_id' column in Transition Sheet.")

    pairs = set()
    for _, row in df_trans.iterrows():
        cid = str(row[company_col]).strip()
        if not cid or cid.lower() == 'nan':
            continue

        if gran_col:
            raw  = str(row[gran_col]).strip().lower()
            gran = 'quarter' if ('quarter' in raw or raw == 'q') else 'annual'
        else:
            gran = 'annual'

        pairs.add((cid, gran))

    result = sorted(pairs, key=lambda x: (x[0], x[1]))
    print(f"  → Extracted {len(result)} unique (company_id, granularity) pairs "
          f"from local Transition Sheet")
    return result


def build_ri_canonical(ri_map_raw):
    canonical = {}
    merged    = defaultdict(list)
    for ri, tables in ri_map_raw.items():
        key = ri.lower().strip()
        if key not in canonical:
            canonical[key] = ri
        merged[canonical[key]].extend(tables)
    return dict(merged)


def detect_year_cols(df_quant, df_initial):
    year_cols = set()
    for df in [df_quant, df_initial]:
        for col in df.columns:
            if col.lower().strip() in NON_YEAR_COLS:
                continue
            if YEAR_PATTERN.match(col.strip()):
                year_cols.add(col.strip())
    return sorted(year_cols)


def year_cols_for_company(df_quant, df_initial, table_ids, all_year_cols):
    q_co = df_quant[df_quant['table_id'].isin(table_ids)]
    i_co = df_initial[df_initial['table_id'].isin(table_ids)]
    active = []
    for y in all_year_cols:
        q_has = y in q_co.columns and q_co[y].replace('', np.nan).dropna().shape[0] > 0
        i_has = y in i_co.columns and i_co[y].replace('', np.nan).dropna().shape[0] > 0
        if q_has or i_has:
            active.append(y)
    return active


# ════════════════════════════════════════════════════════════════
# TRANSITION SHEET PARSER  (per company)
# ════════════════════════════════════════════════════════════════

def parse_transition_for_company(df_trans: pd.DataFrame, company_id: str):
    company_col  = normalise_col(df_trans, ['company_id', 'companyid', 'key'])
    quant_ri_col = normalise_col(df_trans, ['quantri', 'quant_ri', 'ri', 'quantriname'])
    table_id_col = normalise_col(df_trans, ['table_id', 'tableid', 'old table id', 'oldtableid'])
    status_col   = normalise_col(df_trans, ['status', 'processflag', 'process_flag', 'flag', 'pf0', 'pf'])
    period_col   = normalise_col(df_trans, ['period', 'datafrequency', 'data_frequency', 'frequency'])

    for name, col in [('company_id', company_col),
                      ('quant_ri',   quant_ri_col),
                      ('table_id',   table_id_col)]:
        if col is None:
            raise ValueError(f"Cannot find column '{name}' in Transition Sheet. "
                             f"Available: {list(df_trans.columns)}")

    df_co = df_trans[
        df_trans[company_col].str.strip().str.lower() == company_id.strip().lower()
    ]
    if df_co.empty:
        raise ValueError(f"Company '{company_id}' not found in Transition Sheet.")

    ri_map_raw   = defaultdict(list)
    pf0_tables   = set()
    table_period = {}

    for _, row in df_co.iterrows():
        ri  = str(row[quant_ri_col]).strip()
        tbl = str(row[table_id_col]).strip()
        if not ri or not tbl or tbl.lower() in ('nan', ''):
            continue
        ri_map_raw[ri].append(tbl)
        if status_col and str(row[status_col]).strip().lower() in PF0_STATUS_VALUES:
            pf0_tables.add(tbl)
        if period_col:
            table_period[tbl] = str(row[period_col]).strip()

    ri_map = build_ri_canonical(dict(ri_map_raw))
    return ri_map, pf0_tables, table_period, period_col


# ════════════════════════════════════════════════════════════════
# COMPARISON ENGINE
# ════════════════════════════════════════════════════════════════

def vals_differ(a, b):
    if a == b:
        return False
    try:
        return float(a) != float(b)
    except (ValueError, TypeError):
        return a != b


def compare_tables(df_quant, df_initial, table_id, year_cols, pf0_tables):
    errors   = []
    q_rows   = df_quant [df_quant ['table_id'] == table_id].reset_index(drop=True)
    i_rows   = df_initial[df_initial['table_id'] == table_id].reset_index(drop=True)
    q_exists = len(q_rows) > 0
    i_exists = len(i_rows) > 0
    is_pf0   = table_id in pf0_tables

    def err(etype, pk='', detail='', qv='', iv=''):
        return {
            'table_id': table_id, 'primary_key': pk,
            'error_type': etype, 'detail': detail,
            'quant_value': qv, 'initial_value': iv,
            'is_pf0': is_pf0,
        }

    if not q_exists and not i_exists:
        etype  = 'missing_table_in_both_pf0' if is_pf0 else 'missing_table_in_both'
        detail = 'Table absent from both files (PF0)' if is_pf0 else 'Table absent from both files'
        errors.append(err(etype, detail=detail))
        return errors

    if not q_exists:
        errors.append(err('missing_table_in_quant',
                          detail='Present in Initial, absent in Final (Quant)'))
        return errors

    if not i_exists:
        errors.append(err('missing_table_in_initial',
                          detail='Present in Final (Quant), absent in Initial (bkp)'))
        return errors

    q_dict  = {str(r.get('primary_key', '')): r for _, r in q_rows.iterrows()}
    i_dict  = {str(r.get('primary_key', '')): r for _, r in i_rows.iterrows()}
    all_pks = set(q_dict) | set(i_dict)

    for pk in all_pks:
        qr = q_dict.get(pk)
        ir = i_dict.get(pk)
        if qr is None:
            errors.append(err('missing_table_in_quant', pk, 'Row in Initial but not in Final'))
            continue
        if ir is None:
            errors.append(err('missing_table_in_initial', pk, 'Row in Final but not in Initial'))
            continue

        def qv(col): return str(qr.get(col, '')).strip()
        def iv(col): return str(ir.get(col, '')).strip()

        for pg in ['doc_page_num', 'file_page_num']:
            if vals_differ(qv(pg), iv(pg)):
                errors.append(err('change_in_page_number', pk, f'{pg} differs', qv(pg), iv(pg)))
        for y in year_cols:
            if vals_differ(qv(y), iv(y)):
                errors.append(err('change_in_values', pk, f'Year "{y}" differs', qv(y), iv(y)))
        for di in ['dim_1_name', 'dim_2_name', 'dim_3_name', 'dim_4_name',
                   'metric_name', 'source_metric_name']:
            if vals_differ(qv(di), iv(di)):
                errors.append(err('change_in_dim_names', pk, f'{di} differs', qv(di), iv(di)))
        for etype, col in [
            ('change_in_unit',        'unit'),
            ('change_in_unit_type',   'unit_type'),
            ('change_in_note_id',     'note_id'),
            ('change_in_comments',    'comments'),
            ('change_in_indentation', 'indentation'),
        ]:
            if vals_differ(qv(col), iv(col)):
                errors.append(err(etype, pk, f'{col} differs', qv(col), iv(col)))

    return errors


# ════════════════════════════════════════════════════════════════
# COVERAGE CHECK  — per (company_id, granularity) against Azure
# Reports exactly which file is missing: final only, initial only,
# or both. Uses the paths derived from company_id + granularity.
# ════════════════════════════════════════════════════════════════

def _coverage_check_sync(args) -> dict:
    company_id, granularity, n_trans_rows = args

    final_blob_path   = get_final_blob_path(company_id, granularity)
    initial_bkp_prefix = get_initial_bkp_prefix(company_id, granularity)

    final_found   = False
    initial_found = False
    initial_blob  = None
    initial_lm    = None

    # ── Check Final (Quant) file ──────────────────────────────
    try:
        cc = get_container_client()
        cc.get_blob_client(final_blob_path).get_blob_properties()
        final_found = True
    except Exception:
        final_found = False

    # ── Check Initial (bkp) file — scan prefix for latest ────
    try:
        cc = get_container_client()
        latest_blob = None
        latest_time = None
        for blob in cc.list_blobs(name_starts_with=initial_bkp_prefix):
            name = blob.name.lower()
            if '_bkp' in name and (name.endswith('.csv') or name.endswith('.xlsx')):
                if latest_time is None or blob.last_modified > latest_time:
                    latest_time = blob.last_modified
                    latest_blob = blob.name
        if latest_blob:
            initial_found = True
            initial_blob  = latest_blob
            initial_lm    = latest_time
    except Exception:
        initial_found = False

    # ── Determine status — explicit about WHICH file is missing
    if final_found and initial_found:
        status = '✅ Both Found'
    elif final_found and not initial_found:
        status = '⚠️ Initial (bkp) File Missing'
    elif not final_found and initial_found:
        status = '⚠️ Final (Quant) File Missing'
    else:
        status = '❌ Both Files Missing'

    stale = is_blob_stale(initial_lm) if initial_lm else True

    return {
        'company_id':            company_id,
        'granularity':           granularity,
        'transition_rows':       n_trans_rows,
        'final_blob_path':       final_blob_path if final_found else 'NOT FOUND',
        'final_found':           '✅' if final_found else '❌ MISSING',
        'initial_blob_path':     initial_blob if initial_found else f'{initial_bkp_prefix}*_bkp* → NOT FOUND',
        'initial_found':         '✅' if initial_found else '❌ MISSING',
        'initial_last_modified': str(initial_lm) if initial_lm else 'N/A',
        'initial_freshness':     '⚠️ Stale' if stale else '✅ Fresh',
        'status':                status,
    }


async def run_coverage_check_async(pairs: list, df_trans: pd.DataFrame) -> list:
    company_col = normalise_col(df_trans, ['company_id', 'companyid', 'key'])
    gran_col    = normalise_col(df_trans, ['granularity', 'datafrequency', 'data_frequency',
                                            'frequency', 'period', 'period_type', 'annual'])

    def _row_count(cid, gran):
        if company_col is None:
            return 0
        sub = df_trans[df_trans[company_col].str.strip().str.lower() == cid.lower()]
        if gran_col:
            sub = sub[sub[gran_col].str.strip().str.lower().apply(
                lambda v: ('quarter' if ('quarter' in v or v == 'q') else 'annual') == gran
            )]
        return len(sub)

    args_list = [(cid, gran, _row_count(cid, gran)) for cid, gran in pairs]

    loop = asyncio.get_event_loop()
    with concurrent.futures.ThreadPoolExecutor(max_workers=16) as pool:
        tasks   = [loop.run_in_executor(pool, _coverage_check_sync, a) for a in args_list]
        results = await asyncio.gather(*tasks)
    return list(results)


# ════════════════════════════════════════════════════════════════
# ASYNC PER-COMPANY PROCESSOR
# ════════════════════════════════════════════════════════════════

async def process_one_company_async(
    company_id: str,
    granularity: str,
    df_trans: pd.DataFrame,
    semaphore: asyncio.Semaphore,
) -> dict:
    async with semaphore:
        result_base = {
            'company_id':        company_id,
            'granularity':       granularity,
            'ri_summary':        None,
            'detail_rows':       [],
            'missing_rows':      [],
            'year_cols':         [],
            'company_period':    'N/A',
            'bkp_last_modified': None,
            'error':             None,
            'missing_file':      None,   # 'final', 'initial', 'both', or None
        }

        print(f"  ▶ Processing  {company_id} / {granularity}")

        # ── Download final + initial concurrently from Azure ────
        df_quant    = None
        df_initial  = None
        bkp_lm      = None
        final_err   = None
        initial_err = None

        try:
            df_quant = await load_final_file_async(company_id, granularity)
        except Exception as e:
            final_err = str(e)

        try:
            df_initial, bkp_lm = await load_initial_file_async(company_id, granularity)
        except Exception as e:
            initial_err = str(e)

        result_base['bkp_last_modified'] = bkp_lm

        # ── Both missing → skip entirely ──────────────────────
        if final_err and initial_err:
            result_base['error']        = f"Both files missing. Final: {final_err} | Initial: {initial_err}"
            result_base['missing_file'] = 'both'
            print(f"    ❌ SKIP  {company_id}/{granularity} — both files missing")
            return result_base

        # ── Only Final missing ─────────────────────────────────
        if final_err:
            result_base['error']        = f"Final (Quant) file missing: {final_err}"
            result_base['missing_file'] = 'final'
            print(f"    ⚠  SKIP  {company_id}/{granularity} — Final file missing")
            return result_base

        # ── Only Initial missing ───────────────────────────────
        if initial_err:
            result_base['error']        = f"Initial (bkp) file missing: {initial_err}"
            result_base['missing_file'] = 'initial'
            print(f"    ⚠  SKIP  {company_id}/{granularity} — Initial file missing")
            return result_base

        result_base['missing_file'] = None

        if is_blob_stale(bkp_lm):
            print(f"    ⚠  bkp file is stale ({bkp_lm}) — processing anyway")

        # ── Ensure table_id exists in both DataFrames ──────────
        for df, label in [(df_quant, 'Final'), (df_initial, 'Initial')]:
            col = normalise_col(df, ['table_id', 'tableid'])
            if col and col != 'table_id':
                df.rename(columns={col: 'table_id'}, inplace=True)
            if 'table_id' not in df.columns:
                result_base['error'] = f"'table_id' missing in {label} file"
                return result_base

        # ── Detect year columns ────────────────────────────────
        all_year_cols = detect_year_cols(df_quant, df_initial)

        # ── Parse RI map from local transition sheet ───────────
        try:
            ri_map, pf0_tables, table_period, _ = parse_transition_for_company(
                df_trans, company_id
            )
        except ValueError as e:
            result_base['error'] = str(e)
            return result_base

        all_tables   = list({t for tbls in ri_map.values() for t in tbls})
        co_year_cols = year_cols_for_company(df_quant, df_initial, all_tables, all_year_cols)

        # ── Derive period label ────────────────────────────────
        period_col_q = normalise_col(df_quant, ['data_frequency', 'datafrequency',
                                                  'frequency', 'period'])
        company_period = 'N/A'
        if period_col_q and period_col_q in df_quant.columns:
            periods = df_quant[period_col_q].replace('', np.nan).dropna().unique()
            if len(periods):
                company_period = ', '.join(sorted(set(str(p).strip() for p in periods)))

        # ── Compare tables ─────────────────────────────────────
        ri_summary   = defaultdict(lambda: defaultdict(int))
        detail_rows  = []
        missing_rows = []

        for ri, table_ids in ri_map.items():
            unique_tables = list(set(table_ids))
            ri_summary[ri]['total_tables'] += len(unique_tables)

            for tbl in unique_tables:
                is_pf0 = tbl in pf0_tables
                errs   = compare_tables(df_quant, df_initial, tbl, co_year_cols, pf0_tables)

                for e in errs:
                    etype = e['error_type']
                    ri_summary[ri][etype] += 1

                    if etype in MISSING_ERROR_TYPES:
                        ri_summary[ri]['total_missing'] += 1
                        missing_rows.append({'quant_ri': ri, **e})
                    elif is_pf0:
                        ri_summary[ri]['total_pf0_errors'] += 1
                        detail_rows.append({'quant_ri': ri, **e})
                    else:
                        ri_summary[ri]['total_errors'] += 1
                        detail_rows.append({'quant_ri': ri, **e})

        result_base.update({
            'ri_summary':     dict(ri_summary),
            'detail_rows':    detail_rows,
            'missing_rows':   missing_rows,
            'year_cols':      co_year_cols,
            'company_period': company_period,
        })
        print(f"    ✅ Done  {company_id}/{granularity}  "
              f"(RIs={len(ri_map)}, errors={sum(d.get('total_errors', 0) for d in ri_summary.values())})")
        return result_base


# ════════════════════════════════════════════════════════════════
# MAIN ANALYSIS ORCHESTRATOR
# ════════════════════════════════════════════════════════════════

async def run_analysis_async():
    print("=" * 65)
    print("  Quant RI Comparator  —  Local Transition Sheet + Azure Blobs")
    print("=" * 65)

    # Step 1 — Load Transition Sheet from LOCAL path
    df_trans = load_transition_sheet_local()

    # Step 2 — Extract (company_id, granularity) pairs from it
    pairs = extract_company_granularity_pairs(df_trans)
    if not pairs:
        raise ValueError("No (company_id, granularity) pairs found in Transition Sheet.")

    # Step 3 — Process all companies concurrently (files from Azure)
    semaphore     = asyncio.Semaphore(8)
    process_tasks = [
        process_one_company_async(cid, gran, df_trans, semaphore)
        for cid, gran in pairs
    ]

    print(f"\nProcessing {len(pairs)} company/granularity pairs asynchronously...\n")
    company_results = await asyncio.gather(*process_tasks)

    # Step 4 — Coverage check (all companies against Azure)
    print("\nRunning Azure file coverage check...")
    coverage_results = await run_coverage_check_async(pairs, df_trans)

    # Step 5 — Aggregate
    all_ri_summary   = defaultdict(lambda: defaultdict(int))
    all_detail_rows  = []
    all_missing_rows = []
    all_year_cols    = set()
    skipped          = []

    for res in company_results:
        if res['error']:
            skipped.append({
                'company_id':       res['company_id'],
                'granularity':      res['granularity'],
                'missing_file':     res.get('missing_file', 'unknown'),
                'reason':           res['error'],
                'bkp_last_modified': str(res['bkp_last_modified']) if res['bkp_last_modified'] else 'N/A',
            })
            continue

        all_year_cols.update(res['year_cols'])

        for ri, d in (res['ri_summary'] or {}).items():
            for k, v in d.items():
                all_ri_summary[ri][k] += v

        for row in res['detail_rows']:
            all_detail_rows.append({**row, 'company_id': res['company_id'],
                                           'granularity': res['granularity']})
        for row in res['missing_rows']:
            all_missing_rows.append({**row, 'company_id': res['company_id'],
                                            'granularity': res['granularity']})

    return (
        dict(all_ri_summary),
        all_detail_rows,
        all_missing_rows,
        sorted(all_year_cols),
        coverage_results,
        skipped,
        company_results,
    )


# ════════════════════════════════════════════════════════════════
# STYLE HELPERS
# ════════════════════════════════════════════════════════════════

def style_header_row(ws, row, col_start, col_end,
                     fill=HDR_FILL, font=HDR_FONT, height=28):
    ws.row_dimensions[row].height = height
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = CENTER
        cell.border    = BORDER


def style_body_row(ws, row, col_start, col_end, alt=False):
    f = ALT_FILL if alt else WHITE_FILL
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = f
        cell.font      = BODY_FONT
        cell.alignment = LEFT
        cell.border    = BORDER


def set_col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ════════════════════════════════════════════════════════════════
# EXCEL REPORT BUILDER
# ════════════════════════════════════════════════════════════════

def build_report(
    ri_summary,
    detail_rows,
    missing_rows,
    year_cols,
    coverage_results,
    skipped,
    company_results,
    output_path,
):
    wb = Workbook()

    total_errors_all = sum(d.get('total_errors', 0)     for d in ri_summary.values())
    total_pf0_all    = sum(d.get('total_pf0_errors', 0) for d in ri_summary.values())
    total_tables_all = sum(d.get('total_tables', 0)     for d in ri_summary.values())

    # ════════════════════════════════════════
    # SHEET 1 — Summary Dashboard
    # ════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Summary Dashboard'
    n_cols    = 6
    merge_end = get_column_letter(n_cols)

    ws1.merge_cells(f'A1:{merge_end}1')
    ws1['A1'] = '📊  Quant RI Error Rate Report  —  All Companies'
    ws1['A1'].font      = TITLE_FONT
    ws1['A1'].alignment = CENTER
    ws1['A1'].fill      = PatternFill('solid', start_color='EBF3FB')
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells(f'A2:{merge_end}2')
    ws1['A2'] = (
        f'Year Cols: {", ".join(year_cols)}  |  '
        f'Total Tables: {total_tables_all}  |  '
        f'Counted Errors: {total_errors_all}  |  '
        f'PF0 Errors (excl.): {total_pf0_all}  |  '
        f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    )
    ws1['A2'].font      = Font(name='Arial', italic=True, size=9, color='555555')
    ws1['A2'].alignment = CENTER
    ws1.row_dimensions[2].height = 18

    kpis = [
        ('Total Tables',       total_tables_all,  '1F3864'),
        ('Counted Errors',     total_errors_all,  'C00000'),
        ('PF0 Errors (excl.)', total_pf0_all,     '7030A0'),
        ('Quant RIs',          len(ri_summary),   '2E75B6'),
        ('Overall Error Rate',
         f'{round(total_errors_all / total_tables_all * 100, 2) if total_tables_all else 0:.2f}%',
         '375623'),
    ]
    ws1.row_dimensions[3].height = 20
    ws1.row_dimensions[4].height = 32
    for i, (label, val, color) in enumerate(kpis, 1):
        lc = ws1.cell(row=3, column=i, value=label)
        lc.font      = Font(name='Arial', bold=True, size=8, color='FFFFFF')
        lc.fill      = PatternFill('solid', start_color=color)
        lc.alignment = CENTER
        lc.border    = BORDER
        vc = ws1.cell(row=4, column=i, value=val)
        vc.font      = Font(name='Arial', bold=True, size=14, color=color)
        vc.fill      = PatternFill('solid', start_color='F2F2F2')
        vc.alignment = CENTER
        vc.border    = BORDER

    sum_headers = ['Quant RI', 'Total Tables', 'Tables w/ Error',
                   'Error Rate (%)', 'PF0 Errors\n(excl.)', 'Counted Errors']
    for c_idx, h in enumerate(sum_headers, 1):
        ws1.cell(row=6, column=c_idx, value=h)
    style_header_row(ws1, 6, 1, len(sum_headers))

    data_row_start = 7
    dash_row       = data_row_start
    for ri in sorted(ri_summary.keys()):
        d   = ri_summary[ri]
        tt  = d.get('total_tables', 0)
        te  = d.get('total_errors', 0)
        tpf = d.get('total_pf0_errors', 0)
        er  = round(te / tt * 100, 2) if tt else 0.0
        twe = len(set(rd['table_id'] for rd in detail_rows
                      if rd['quant_ri'] == ri and not rd.get('is_pf0', False)))
        for c_idx, val in enumerate([ri, tt, twe, er, tpf, te], 1):
            ws1.cell(row=dash_row, column=c_idx, value=val)
        style_body_row(ws1, dash_row, 1, len(sum_headers), alt=(dash_row % 2 == 0))
        rc = ws1.cell(row=dash_row, column=4)
        rc.fill          = OK_FILL if er == 0 else (WARN_FILL if er < 20 else ERR_FILL)
        rc.font          = BOLD_FONT
        rc.number_format = '0.00"%"'
        dash_row += 1

    last_data = dash_row - 1
    ws1.cell(row=dash_row, column=1, value='TOTAL')
    ws1.cell(row=dash_row, column=2, value=f'=SUM(B{data_row_start}:B{last_data})')
    ws1.cell(row=dash_row, column=3, value=f'=SUM(C{data_row_start}:C{last_data})')
    ws1.cell(row=dash_row, column=4, value=f'=IFERROR(F{dash_row}/B{dash_row}*100,0)')
    ws1.cell(row=dash_row, column=4).number_format = '0.00"%"'
    ws1.cell(row=dash_row, column=5, value=f'=SUM(E{data_row_start}:E{last_data})')
    ws1.cell(row=dash_row, column=6, value=f'=SUM(F{data_row_start}:F{last_data})')
    style_header_row(ws1, dash_row, 1, len(sum_headers), fill=SUB_FILL, height=22)
    set_col_widths(ws1, {'A': 26, 'B': 14, 'C': 16, 'D': 16, 'E': 18, 'F': 16})
    ws1.freeze_panes = 'A7'

    # ════════════════════════════════════════
    # SHEET 2 — Per-Company Summary
    # ════════════════════════════════════════
    ws_co = wb.create_sheet('Company Summary')
    ws_co.merge_cells('A1:H1')
    ws_co['A1'] = 'Per-Company Error Summary'
    ws_co['A1'].font      = TITLE_FONT
    ws_co['A1'].alignment = CENTER
    ws_co['A1'].fill      = PatternFill('solid', start_color='EBF3FB')
    ws_co.row_dimensions[1].height = 30

    co_h = ['Company ID', 'Granularity', 'Period', 'Total Tables',
             'Tables w/ Error', 'Error Rate (%)', 'Counted Errors', 'PF0 Errors (excl.)']
    for c_idx, h in enumerate(co_h, 1):
        ws_co.cell(row=2, column=c_idx, value=h)
    style_header_row(ws_co, 2, 1, len(co_h))

    for r_idx, res in enumerate(
        sorted(company_results, key=lambda x: (x['company_id'], x['granularity'])), 3
    ):
        if res['error']:
            continue
        ri_s  = res['ri_summary'] or {}
        tt    = sum(d.get('total_tables', 0) for d in ri_s.values())
        te    = sum(d.get('total_errors', 0) for d in ri_s.values())
        tpf   = sum(d.get('total_pf0_errors', 0) for d in ri_s.values())
        er    = round(te / tt * 100, 2) if tt else 0.0
        twe   = len(set(rd['table_id'] for rd in res['detail_rows'] if not rd.get('is_pf0', False)))
        for c_idx, v in enumerate(
            [res['company_id'], res['granularity'], res['company_period'],
             tt, twe, er, te, tpf], 1
        ):
            ws_co.cell(row=r_idx, column=c_idx, value=v)
        style_body_row(ws_co, r_idx, 1, len(co_h), alt=(r_idx % 2 == 0))
        rc = ws_co.cell(row=r_idx, column=6)
        rc.fill          = OK_FILL if er == 0 else (WARN_FILL if er < 20 else ERR_FILL)
        rc.font          = BOLD_FONT
        rc.number_format = '0.00"%"'
    set_col_widths(ws_co, {'A': 22, 'B': 12, 'C': 16, 'D': 14, 'E': 16, 'F': 16, 'G': 16, 'H': 18})
    ws_co.freeze_panes = 'A3'

    # ════════════════════════════════════════
    # SHEET 3 — RI Error Rate
    # ════════════════════════════════════════
    ws2 = wb.create_sheet('RI Error Rate')
    ws2.merge_cells(f'A1:{get_column_letter(4 + len(COUNTED_ERROR_TYPES))}1')
    ws2['A1'] = 'Error Rate by Quant RI  —  All Companies'
    ws2['A1'].font      = TITLE_FONT
    ws2['A1'].alignment = CENTER
    ws2['A1'].fill      = PatternFill('solid', start_color='EBF3FB')
    ws2.row_dimensions[1].height = 30

    ri_h = (['Quant RI', 'Total Tables', 'Tables w/ Error', 'Error Rate (%)']
             + [ERROR_LABELS[e] for e in COUNTED_ERROR_TYPES])
    for c_idx, h in enumerate(ri_h, 1):
        ws2.cell(row=2, column=c_idx, value=h)
    style_header_row(ws2, 2, 1, len(ri_h))

    for r_idx, ri in enumerate(sorted(ri_summary.keys()), 3):
        d   = ri_summary[ri]
        tt  = d.get('total_tables', 0)
        te  = d.get('total_errors', 0)
        er  = round(te / tt * 100, 2) if tt else 0.0
        twe = len(set(rd['table_id'] for rd in detail_rows
                      if rd['quant_ri'] == ri and not rd.get('is_pf0', False)))
        for c_idx, v in enumerate(
            [ri, tt, twe, er] + [d.get(e, 0) for e in COUNTED_ERROR_TYPES], 1
        ):
            ws2.cell(row=r_idx, column=c_idx, value=v)
        style_body_row(ws2, r_idx, 1, len(ri_h), alt=(r_idx % 2 == 0))
        rc = ws2.cell(row=r_idx, column=4)
        rc.fill          = OK_FILL if er == 0 else (WARN_FILL if er < 20 else ERR_FILL)
        rc.font          = BOLD_FONT
        rc.number_format = '0.00"%"'
    w2w = {'A': 26, 'B': 14, 'C': 16, 'D': 16}
    for i in range(5, 5 + len(COUNTED_ERROR_TYPES)):
        w2w[get_column_letter(i)] = 20
    set_col_widths(ws2, w2w)
    ws2.freeze_panes = 'A3'

    # ════════════════════════════════════════
    # SHEET 4 — Error Types by RI (counted only — no missing rows)
    # ════════════════════════════════════════
    ws4 = wb.create_sheet('Error Types by RI')
    ws4.merge_cells('A1:I1')
    ws4['A1'] = 'Error Types by Quant RI  —  Counted Errors Only'
    ws4['A1'].font      = TITLE_FONT
    ws4['A1'].alignment = CENTER
    ws4['A1'].fill      = PatternFill('solid', start_color='EBF3FB')
    ws4.row_dimensions[1].height = 30

    et_h = ['Company ID', 'Granularity', 'Quant RI', 'Table ID', 'Primary Key',
            'Error Type', 'Detail', 'Final Value', 'Initial Value']
    for c_idx, h in enumerate(et_h, 1):
        ws4.cell(row=2, column=c_idx, value=h)
    style_header_row(ws4, 2, 1, len(et_h))

    for r_idx, rd in enumerate(
        sorted(detail_rows,
               key=lambda x: (x.get('company_id', ''), x['quant_ri'].lower(), x['error_type'])),
        3
    ):
        for c_idx, v in enumerate([
            rd.get('company_id', ''), rd.get('granularity', ''),
            rd['quant_ri'], rd['table_id'], rd['primary_key'],
            ERROR_LABELS.get(rd['error_type'], rd['error_type']),
            rd['detail'], rd['quant_value'], rd['initial_value'],
        ], 1):
            ws4.cell(row=r_idx, column=c_idx, value=v)
        style_body_row(ws4, r_idx, 1, len(et_h), alt=(r_idx % 2 == 0))
        ws4.cell(row=r_idx, column=6).fill = ERR_FILL
        ws4.cell(row=r_idx, column=6).font = BOLD_FONT

    if not detail_rows:
        ws4.merge_cells(f'A3:{get_column_letter(len(et_h))}3')
        ws4.cell(row=3, column=1, value='No counted errors found.')
        ws4.cell(row=3, column=1).font      = Font(name='Arial', italic=True, size=9, color='888888')
        ws4.cell(row=3, column=1).alignment = CENTER

    set_col_widths(ws4, {'A': 20, 'B': 12, 'C': 26, 'D': 22, 'E': 28,
                         'F': 28, 'G': 44, 'H': 22, 'I': 22})
    ws4.freeze_panes = 'A3'

    # ════════════════════════════════════════
    # SHEET 5 — Missing Tables (separate, not in error rate)
    # ════════════════════════════════════════
    ws6 = wb.create_sheet('Missing Tables')
    ws6.merge_cells('A1:G1')
    ws6['A1'] = 'Missing Tables  —  All Companies  (excluded from error rate & all other sheets)'
    ws6['A1'].font      = TITLE_FONT
    ws6['A1'].alignment = CENTER
    ws6['A1'].fill      = PatternFill('solid', start_color='EBF3FB')
    ws6.row_dimensions[1].height = 30

    ws6.merge_cells('A2:G2')
    ws6['A2'] = ('🔴 Absent from Both Files  |  '
                 '🟣 Absent from Both Files (PF0)  |  '
                 '🟡 Present in Final Only (Not in Initial)')
    ws6['A2'].font      = Font(name='Arial', italic=True, size=9, color='444444')
    ws6['A2'].alignment = CENTER
    ws6.row_dimensions[2].height = 16

    miss_h = ['Company ID', 'Granularity', 'Quant RI', 'Table ID',
              'Missing Type', 'Detail', 'PF0?']
    for c_idx, h in enumerate(miss_h, 1):
        ws6.cell(row=3, column=c_idx, value=h)
    style_header_row(ws6, 3, 1, len(miss_h))

    type_order = {
        'missing_table_in_both_pf0': 0,
        'missing_table_in_both':     1,
        'missing_table_in_initial':  2,
    }
    for r_idx, rd in enumerate(
        sorted(missing_rows,
               key=lambda x: (type_order.get(x['error_type'], 9),
                               x.get('company_id', ''), x['quant_ri'].lower())),
        4
    ):
        etype = rd['error_type']
        for c_idx, v in enumerate([
            rd.get('company_id', ''), rd.get('granularity', ''),
            rd['quant_ri'], rd['table_id'],
            MISSING_TYPE_LABELS.get(etype, etype),
            rd['detail'],
            '✅ PF0' if rd.get('is_pf0') else '',
        ], 1):
            ws6.cell(row=r_idx, column=c_idx, value=v)
        style_body_row(ws6, r_idx, 1, len(miss_h), alt=(r_idx % 2 == 0))
        mc      = ws6.cell(row=r_idx, column=5)
        mc.font = BOLD_FONT
        mc.fill = (
            PatternFill('solid', start_color='E8D5F5') if etype == 'missing_table_in_both_pf0' else
            PatternFill('solid', start_color='FFE0E0') if etype == 'missing_table_in_both'     else
            PatternFill('solid', start_color='FFF2CC')
        )

    if not missing_rows:
        ws6.merge_cells(f'A4:{get_column_letter(len(miss_h))}4')
        ws6.cell(row=4, column=1, value='No missing tables found.')
        ws6.cell(row=4, column=1).font      = Font(name='Arial', italic=True, size=9, color='888888')
        ws6.cell(row=4, column=1).alignment = CENTER

    set_col_widths(ws6, {'A': 22, 'B': 12, 'C': 26, 'D': 28, 'E': 34, 'F': 50, 'G': 10})
    ws6.freeze_panes = 'A4'

    # ════════════════════════════════════════
    # SHEET 6 — Azure Coverage
    # Shows exactly which file is missing per company+granularity
    # ════════════════════════════════════════
    ws5 = wb.create_sheet('Azure Coverage')
    ws5.merge_cells('A1:J1')
    ws5['A1'] = ('Azure File Coverage  —  All Companies from Local Transition Sheet  '
                 '(paths derived from company_id + granularity)')
    ws5['A1'].font      = TITLE_FONT
    ws5['A1'].alignment = CENTER
    ws5['A1'].fill      = PatternFill('solid', start_color='EBF3FB')
    ws5.row_dimensions[1].height = 30

    ws5.merge_cells('A2:J2')
    ws5['A2'] = ('✅ Both Found  |  '
                 '⚠️ Final (Quant) File Missing  |  '
                 '⚠️ Initial (bkp) File Missing  |  '
                 '❌ Both Files Missing')
    ws5['A2'].font      = Font(name='Arial', italic=True, size=9, color='444444')
    ws5['A2'].alignment = CENTER
    ws5.row_dimensions[2].height = 16

    cov_h = [
        'Company ID', 'Granularity', 'Transition Rows',
        'Final (Quant) Path', 'Final Found?',
        'Initial (bkp) Path', 'Initial Found?',
        'BKP Last Modified', 'BKP Freshness',
        'Status',
    ]
    for c_idx, h in enumerate(cov_h, 1):
        ws5.cell(row=3, column=c_idx, value=h)
    style_header_row(ws5, 3, 1, len(cov_h))

    # Sort: both-missing first, then partial, then ok
    coverage_results.sort(key=lambda r: (
        0 if '❌' in r['status'] else 1 if '⚠️' in r['status'] else 2,
        r['company_id'].lower(), r['granularity']
    ))

    for r_idx, cr in enumerate(coverage_results, 4):
        for c_idx, v in enumerate([
            cr['company_id'],
            cr['granularity'],
            cr['transition_rows'],
            cr['final_blob_path'],
            cr['final_found'],
            cr['initial_blob_path'],
            cr['initial_found'],
            cr['initial_last_modified'],
            cr['initial_freshness'],
            cr['status'],
        ], 1):
            ws5.cell(row=r_idx, column=c_idx, value=v)
        style_body_row(ws5, r_idx, 1, len(cov_h), alt=(r_idx % 2 == 0))

        # Color Final Found cell
        fc = ws5.cell(row=r_idx, column=5)
        fc.fill = OK_FILL if cr['final_found'] == '✅' else ERR_FILL
        fc.font = BOLD_FONT

        # Color Initial Found cell
        ic = ws5.cell(row=r_idx, column=7)
        ic.fill = OK_FILL if cr['initial_found'] == '✅' else ERR_FILL
        ic.font = BOLD_FONT

        # Color Status cell
        sc = ws5.cell(row=r_idx, column=10)
        sc.fill = (OK_FILL   if '✅' in cr['status'] else
                   WARN_FILL if '⚠️' in cr['status'] else ERR_FILL)
        sc.font = BOLD_FONT

    set_col_widths(ws5, {
        'A': 22, 'B': 12, 'C': 16,
        'D': 65, 'E': 14,
        'F': 65, 'G': 14,
        'H': 26, 'I': 14, 'J': 30,
    })
    ws5.freeze_panes = 'A4'

    # ════════════════════════════════════════
    # SHEET 7 — Skipped Companies
    # Shows which file was missing (final / initial / both)
    # ════════════════════════════════════════
    ws_sk = wb.create_sheet('Skipped')
    ws_sk.merge_cells('A1:E1')
    ws_sk['A1'] = 'Skipped Companies  —  File Missing or Download Error'
    ws_sk['A1'].font      = TITLE_FONT
    ws_sk['A1'].alignment = CENTER
    ws_sk['A1'].fill      = PatternFill('solid', start_color='EBF3FB')
    ws_sk.row_dimensions[1].height = 30

    sk_h = ['Company ID', 'Granularity', 'Missing File', 'BKP Last Modified', 'Reason']
    for c_idx, h in enumerate(sk_h, 1):
        ws_sk.cell(row=2, column=c_idx, value=h)
    style_header_row(ws_sk, 2, 1, len(sk_h))

    for r_idx, sk in enumerate(skipped, 3):
        miss = sk.get('missing_file', 'unknown')
        miss_label = {
            'both':    '❌ Both Files Missing',
            'final':   '⚠️ Final (Quant) Missing',
            'initial': '⚠️ Initial (bkp) Missing',
        }.get(miss, f'⚠️ {miss}')

        for c_idx, v in enumerate([
            sk['company_id'],
            sk['granularity'],
            miss_label,
            sk['bkp_last_modified'],
            sk['reason'],
        ], 1):
            ws_sk.cell(row=r_idx, column=c_idx, value=v)
        style_body_row(ws_sk, r_idx, 1, len(sk_h), alt=(r_idx % 2 == 0))

        mc = ws_sk.cell(row=r_idx, column=3)
        mc.font = BOLD_FONT
        mc.fill = ERR_FILL if '❌' in miss_label else WARN_FILL
        ws_sk.cell(row=r_idx, column=5).fill = ERR_FILL

    if not skipped:
        ws_sk.merge_cells(f'A3:{get_column_letter(len(sk_h))}3')
        ws_sk.cell(row=3, column=1, value='No companies were skipped.')
        ws_sk.cell(row=3, column=1).font      = Font(name='Arial', italic=True, size=9, color='888888')
        ws_sk.cell(row=3, column=1).alignment = CENTER

    set_col_widths(ws_sk, {'A': 22, 'B': 12, 'C': 28, 'D': 26, 'E': 60})
    ws_sk.freeze_panes = 'A3'

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"\n✅  Report saved → {output_path}")


# ════════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print(f"Transition (local) : {TRANSITION_LOCAL_PATH}")
    print(f"Output             : {OUTPUT_PATH}\n")

    (
        ri_summary,
        detail_rows,
        missing_rows,
        year_cols,
        coverage_results,
        skipped,
        company_results,
    ) = asyncio.run(run_analysis_async())

    build_report(
        ri_summary       = ri_summary,
        detail_rows      = detail_rows,
        missing_rows     = missing_rows,
        year_cols        = year_cols,
        coverage_results = coverage_results,
        skipped          = skipped,
        company_results  = company_results,
        output_path      = OUTPUT_PATH,
    )